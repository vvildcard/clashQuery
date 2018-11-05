# python 3.7
# A script to query the Clash Royale API and detect inactive players in your clan. 
# by vvildcard
# 2018/11/01
#
# Notes:
#   You MUST have a worksheet named 'clanDonations.xlsx' in the script's working directory.
#   The script only reads the first worksheet.
#   The first worksheet must be named as the clan's ID.
#
# To Do:
#   Test a worksheet named 'Sheet1'
#   Add other indicators of activity (war participation, etc)
#   Conditional formatting for donation cells


# -----------------
# Import modules // Define functions // Set variables
# -----------------

import requests, datetime, os
from openpyxl import load_workbook


def getClanID():
    tempClanID = input(bcolors.BOLD + "Input your Clan ID: " + bcolors.ENDC)
    while True:
        if tempClanID[0] == "#":  # Remove the #
            tempClanID = tempClanID[1:]
        try:
            clanIDSearch = requests.request("GET", url + "clans/%23" + tempClanID, headers=headers).json()
            print("Found clan: " + str(clanIDSearch['name']))
            break
        except KeyError:
            tempClanID = input(
                bcolors.WARNING + "Clan not found.\n" +
                bcolors.BOLD + "Try again: " + bcolors.ENDC)
    return tempClanID


def getToken():
    while True:
        getTokenToken = ''  # Start with a blank variable (so we can attempt to get it from token.txt)
        try:
            tokenFile = open('token.txt', 'r')
            if tokenFile.read() == '':
                askForToken("Invalid")
            if tokenTest(tokenFile):
                getTokenToken = tokenFile
                tokenFile.seek(0)
                break  # The token was valid
            else:
                tokenFile.close()                    # Close the file
                blank = ''                           # Clear the token so we can try again.
                tokenFile = open('token.txt', 'w+')  # Open token.txt
                tokenFile.write(blank)               # Clear the file
                tokenFile.close()                    # Close the file
        except FileNotFoundError:  # This always means token.txt doesn't exist.
            tokenInput = askForToken("Missing")
            testedToken = tokenTest(tokenInput)
            if testedToken:
                getTokenToken = tokenInput
            else:
                getTokenToken = ''  # Clear the token and try again.
            tokenFile = open('token.txt', 'w+')  # Create/open token.txt
            tokenFile.write(getTokenToken)  # Write the token into the file
            tokenFile.close()
    return getTokenToken


def askForToken(note):
    tokenInput = input(
        bcolors.WARNING + note + " token.\n" +
        bcolors.ENDC + "If you don't have a token, get one here: " + bcolors.OKBLUE + "developer.clashroyale.com\n" +
        bcolors.FAIL + "Your token will be stored in plain-text here:\n" +
        bcolors.ENDC + os.getcwd() + '\\token.txt' + "\n" +
        bcolors.BOLD + "\nInput your token: " + bcolors.ENDC)
    while tokenInput == '':  # Ask again if the token wasn't given
        tokenInput = input(bcolors.BOLD + "Input your token: " + bcolors.ENDC)
    tokenFile = open('token.txt', 'w+')  # Open token.txt
    tokenFile.write(tokenInput)  # Write the user response to the file
    tokenFile.close()  # Close the file
    return tokenInput


def tokenTest(tempTokenTest):
    try:  # Create a header
        tempTokenTest.seek(0)  # Make sure to read the whole file
        tokenTestHeaders = {"Authorization": "Bearer: " + tempTokenTest.read()}  # Use the token file if it exists
    except AttributeError:
        tokenTestHeaders = {"Authorization": "Bearer: " + tempTokenTest}  # Use the string
    try:
        tokenTestURL = "https://api.clashroyale.com/v1/cards"
        tokenTestRequest = requests.request("GET", tokenTestURL, headers=tokenTestHeaders)  # Test the token
        tokenTestRequest.raise_for_status()
    except requests.exceptions.HTTPError as errh:
        print("HTTP Error:", errh)
        return False
    except requests.exceptions.ConnectionError as errc:
        print("Error Connecting:", errc)
        return False
    except requests.exceptions.Timeout as errt:
        print("Timeout Error:", errt)
        return False
    except requests.exceptions.RequestException as err:
        print("Error!", err)
        return False
    return True


def createWorkbook():
    from openpyxl import Workbook
    wb = Workbook()  # Create a workbook
    ws = wb.active  # Set the active sheet
    ws.title = getClanID()  # Set the Sheet name to the Clan ID
    ws['A1'] = "Name"
    ws['B1'] = "Role"
    ws['C1'] = "LastSeen"
    wb.save("clanDonations.xlsx")
    return True

# Text Colors
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
# Sample usage:
# print bcolors.WARNING + "Warning: Something broke. Continue?"
#      + bcolors.ENDC

date = datetime.datetime.now()
todayDate = date.strftime("%Y" + "-" + "%m" + "-" + "%d")
url = "https://api.clashroyale.com/v1/"
token = getToken()
headers = {"Authorization": "Bearer: " + token.read()}

# -----------------
# Load the data
# -----------------

# Load the Spreadsheet
while True:
    try:
        wb = load_workbook("clanDonations.xlsx", data_only=True)
        ws = wb.active  # Set the active sheet
        if ws.title[0:5] == "Sheet":
            ws.title = getClanID()  # Set the Sheet name to the Clan ID
        break  # Workbook found
    except FileNotFoundError:
        cwd = os.getcwd()  # Current working directory
        print(bcolors.WARNING + "Missing workbook.\n" + bcolors.ENDC +
              bcolors.FAIL + "If you already have a workbook, place it here:\n" +
              bcolors.ENDC + str(cwd))
        createWorkbook()
clanID = wb.sheetnames[0]  # Get the clan ID from the Sheet Name
ws = wb[clanID]  # Set the active sheet

# Get clan info from the API
clanMembers = requests.request("GET", url + "clans/%23" + clanID + "/members", headers=headers).json()
# print(clanMembers)


# -----------------
# Parse and prep the data
# -----------------

# Add a new column for today. Skip if today's date is already recorded.
if ws["D1"].value != todayDate:
    ws.insert_cols(4)
    ws["D1"] = todayDate

#  Make a dictionary
#  like this: [{'member1': ['elder', '2018-10-28', '100']}, {'member2': ['member', '2018-10-26', '0']}]
tempDict = {}
for member in clanMembers['items']:
    tempDict[member['name']] = [member['role'], todayDate, member['donations']]
# print(tempDict); print("\n\n")

memberList = []
for i in range(1, ws.max_row):
    memberList.append(str(ws.cell(row=i + 1, column=1).value))
# print(memberList)

# Add new members
for member in tempDict:
    # print(member)
    if member not in memberList:
        ws.cell(row=ws.max_row + 1, column=1).value = str(member)

# -----------------
# Merge the data
# -----------------

# Search for each member and update their role, last seen date and donations
for member in tempDict:
    # print("member: " + str(member) + "; data: " + str(tempDict[member]))
    for i in range(1, ws.max_row):
        # print("Search: " + str(ws.cell(row=i, column=1).value))
        i += 1
        if str(ws.cell(row=i, column=1).value) == str(member):
            # print(ws.cell(row=i, column=1).value)
            ws.cell(row=i, column=2).value = tempDict[member][0]  # role
            ws.cell(row=i, column=3).value = tempDict[member][1]  # lastSeen
            ws.cell(row=i, column=4).value = tempDict[member][2]  # donations

token.close()
wb.save("clanDonations.xlsx")
wb.close()
